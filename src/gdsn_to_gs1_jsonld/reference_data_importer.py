"""Offline import helpers for public reference source data."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_GDSN_XLSX = Path(
    "reference_data/raw_public/"
    "GDSN_Attributes_with_BMSId_xPath_3.1.36_June_5_2026.xlsx"
)
DEFAULT_WEBVOC = Path("webvoc/current/gs1Voc.jsonld")
DEFAULT_SOURCE_MANIFEST = Path("reference_data/source_manifest.json")
DEFAULT_OUTPUT_DIR = Path("reference_data/normalized")

GDSN_SOURCE_VERSION = "3.1.36"
WEBVOC_SOURCE_VERSION = "1.17"

GDSN_OUTPUT_BASE = "gdsn_attributes_bms_xpath_3_1_36"
WEBVOC_PROPERTIES_OUTPUT_BASE = "webvoc_properties_1_17"
WEBVOC_CLASSES_OUTPUT_BASE = "webvoc_classes_1_17"

GDSN_FIELDS = (
    "bms_id",
    "message",
    "sunrise",
    "sunset",
    "change_type",
    "xpath",
    "module",
    "row_type",
    "parent_class",
    "attribute_name",
    "multiplicity",
    "length",
    "data_type",
    "named_association",
    "class_associated_to",
    "code_list_enumeration",
    "code_list_name",
    "bms_code_list_id",
    "language_enabled",
    "uom_enabled",
    "currency_enabled",
    "global_local",
    "semantic_resource_urn",
    "definition",
    "source_sheet",
    "source_version",
    "is_deleted",
    "is_candidate_source",
)

WEBVOC_PROPERTY_FIELDS = (
    "term_id",
    "compact_name",
    "label",
    "comment",
    "domain",
    "range",
    "sub_property_of",
    "type",
    "is_link_type",
    "term_status",
    "version",
    "last_modified",
)

WEBVOC_CLASS_FIELDS = (
    "term_id",
    "compact_name",
    "label",
    "comment",
    "sub_class_of",
    "type",
    "term_status",
    "version",
    "last_modified",
)

SOURCE_MANIFEST_REQUIRED_FIELDS = (
    "source_id",
    "title",
    "source_url",
    "retrieved_at",
    "version",
    "local_path",
    "sha256",
    "public_accessible",
    "authoritative_or_derived",
    "used_by",
    "license_or_rights_note",
    "usage_note",
)

_HEADER_FIELD_MAP = {
    "bmsid": "bms_id",
    "message": "message",
    "sunrise": "sunrise",
    "sunset": "sunset",
    "changetype": "change_type",
    "xpath": "xpath",
    "module": "module",
    "type": "row_type",
    "parentclass": "parent_class",
    "name": "attribute_name",
    "multiplicity": "multiplicity",
    "length": "length",
    "datatype": "data_type",
    "namedassociation": "named_association",
    "classassociatedto": "class_associated_to",
    "codelistenumeration": "code_list_enumeration",
    "codelistname": "code_list_name",
    "bmscodelistid": "bms_code_list_id",
    "languageenabled": "language_enabled",
    "uomenabled": "uom_enabled",
    "currencyenabled": "currency_enabled",
    "globallocal": "global_local",
    "semanticresourceurn": "semantic_resource_urn",
    "definition": "definition",
}


@dataclass(frozen=True)
class GDSNImportResult:
    rows: list[dict[str, Any]]
    workbook_sheet_count: int
    selected_sheet: str
    active_row_count: int
    deleted_row_count: int


@dataclass(frozen=True)
class ReferenceDataImport:
    source_manifest: dict[str, Any]
    gdsn_rows: list[dict[str, Any]]
    webvoc_properties: list[dict[str, Any]]
    webvoc_classes: list[dict[str, Any]]
    summary: dict[str, Any]


def sha256_file(path: str | Path) -> str:
    """Return the SHA-256 checksum for a local file."""
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_source_manifest(path: str | Path = DEFAULT_SOURCE_MANIFEST) -> dict[str, Any]:
    """Load and validate the source manifest structure."""
    manifest = json.loads(Path(path).read_text(encoding="utf-8-sig"))
    if not isinstance(manifest, dict):
        raise ValueError("Source manifest must be a JSON object.")
    sources = manifest.get("sources")
    if not isinstance(sources, list) or not sources:
        raise ValueError("Source manifest must contain a non-empty sources list.")
    for index, source in enumerate(sources, start=1):
        if not isinstance(source, dict):
            raise ValueError(f"Manifest source {index} must be an object.")
        missing = [
            field for field in SOURCE_MANIFEST_REQUIRED_FIELDS if field not in source
        ]
        if missing:
            raise ValueError(
                f"Manifest source {source.get('source_id', index)!r} is missing "
                f"required field(s): {', '.join(missing)}"
            )
    return manifest


def validate_manifest_checksums(
    manifest: dict[str, Any],
    *,
    base_dir: str | Path = ".",
) -> list[dict[str, Any]]:
    """Compare manifest checksums with files that are available locally."""
    base_path = Path(base_dir)
    checks: list[dict[str, Any]] = []
    for source in manifest["sources"]:
        local_path = base_path / source["local_path"]
        actual = sha256_file(local_path) if local_path.is_file() else ""
        expected = str(source["sha256"]).lower()
        checks.append(
            {
                "source_id": source["source_id"],
                "local_path": source["local_path"],
                "expected_sha256": expected,
                "actual_sha256": actual,
                "exists": local_path.is_file(),
                "matches": bool(actual) and actual == expected,
            }
        )
    return checks


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time().isoformat() == "00:00:00":
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else f"{value:.15g}"
    return str(value).replace("\r\n", "\n").strip()


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _cell_to_text(value).lower())


def _field_for_header(value: Any) -> str:
    return _HEADER_FIELD_MAP.get(_header_key(value), "")


def _row_type(row: dict[str, Any]) -> str:
    return str(row.get("row_type", "")).strip().lower()


def _is_candidate_source(row: dict[str, Any]) -> bool:
    return bool(
        _row_type(row) == "attribute"
        and not row["is_deleted"]
        and row.get("bms_id")
        and row.get("xpath")
        and row.get("attribute_name")
    )


def _normalize_gdsn_row(
    headers: list[Any],
    values: Iterable[Any],
    *,
    sheet_name: str,
    source_version: str,
    is_deleted: bool,
) -> dict[str, Any] | None:
    normalized: dict[str, Any] = {field: "" for field in GDSN_FIELDS}
    for header, value in zip(headers, values):
        field = _field_for_header(header)
        if field:
            normalized[field] = _cell_to_text(value)
    if not any(normalized[field] for field in GDSN_FIELDS[:24]):
        return None
    normalized["source_sheet"] = sheet_name
    normalized["source_version"] = source_version
    normalized["is_deleted"] = is_deleted
    if is_deleted and not normalized["change_type"]:
        normalized["change_type"] = "D"
    normalized["is_candidate_source"] = _is_candidate_source(normalized)
    return normalized


def _normalize_gdsn_sheet(
    worksheet: Any,
    *,
    source_version: str,
    is_deleted: bool,
) -> list[dict[str, Any]]:
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        return []
    normalized: list[dict[str, Any]] = []
    for values in rows:
        row = _normalize_gdsn_row(
            headers,
            values,
            sheet_name=worksheet.title,
            source_version=source_version,
            is_deleted=is_deleted,
        )
        if row:
            normalized.append(row)
    return normalized


def load_gdsn_bms_xpath_workbook(
    path: str | Path,
    *,
    source_version: str = GDSN_SOURCE_VERSION,
) -> GDSNImportResult:
    """Load and normalize the public GDSN BMS/XPath workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if source_version in workbook.sheetnames:
        selected_sheet = source_version
    else:
        version_sheets = [
            name for name in workbook.sheetnames if name.strip().startswith(source_version)
        ]
        if not version_sheets:
            raise ValueError(
                f"Workbook does not contain expected version sheet {source_version!r}."
            )
        selected_sheet = version_sheets[0]

    active_rows = _normalize_gdsn_sheet(
        workbook[selected_sheet],
        source_version=source_version,
        is_deleted=False,
    )
    deleted_rows: list[dict[str, Any]] = []
    if "Deleted Attributes" in workbook.sheetnames:
        deleted_rows = _normalize_gdsn_sheet(
            workbook["Deleted Attributes"],
            source_version=source_version,
            is_deleted=True,
        )
    return GDSNImportResult(
        rows=[*active_rows, *deleted_rows],
        workbook_sheet_count=len(workbook.sheetnames),
        selected_sheet=selected_sheet,
        active_row_count=len(active_rows),
        deleted_row_count=len(deleted_rows),
    )


def load_webvoc_jsonld(path: str | Path = DEFAULT_WEBVOC) -> dict[str, Any]:
    """Load a Web Vocabulary JSON-LD file, including UTF-8 BOM variants."""
    return json.loads(Path(path).read_text(encoding="utf-8-sig"))


def _graph(data: dict[str, Any]) -> list[dict[str, Any]]:
    graph = data.get("@graph", [])
    if not isinstance(graph, list):
        return []
    return [item for item in graph if isinstance(item, dict)]


def _types(item: dict[str, Any]) -> list[str]:
    value = item.get("@type")
    if value is None:
        return []
    if isinstance(value, list):
        return [str(entry) for entry in value]
    return [str(value)]


def _values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        values: list[str] = []
        for entry in value:
            values.extend(_values(entry))
        return values
    if isinstance(value, dict):
        if "@id" in value:
            return [str(value["@id"])]
        if "@value" in value:
            return [str(value["@value"])]
        if "@list" in value:
            return _values(value["@list"])
        return []
    return [str(value)]


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        if "@value" in value:
            return str(value["@value"])
        if "@id" in value:
            return str(value["@id"])
        if "@list" in value:
            return "; ".join(_values(value["@list"]))
        return ""
    if isinstance(value, list):
        return "; ".join(item for item in (_text(entry) for entry in value) if item)
    return str(value)


def _compact_name(term_id: str) -> str:
    return term_id.split(":", 1)[1] if ":" in term_id else term_id


def _is_class(item: dict[str, Any]) -> bool:
    return bool({"owl:Class", "rdfs:Class"} & set(_types(item)))


def _is_property(item: dict[str, Any]) -> bool:
    return bool(
        {"rdf:Property", "owl:ObjectProperty", "owl:DatatypeProperty"}
        & set(_types(item))
    )


def _load_linktypes_for_webvoc(webvoc_path: str | Path) -> dict[str, Any]:
    linktypes_path = Path(webvoc_path).with_name("linktypes.json")
    if not linktypes_path.is_file():
        return {}
    loaded = json.loads(linktypes_path.read_text(encoding="utf-8-sig"))
    return loaded if isinstance(loaded, dict) else {}


def extract_webvoc_classes(
    data: dict[str, Any],
    *,
    version: str = WEBVOC_SOURCE_VERSION,
    last_modified: str = "",
) -> list[dict[str, Any]]:
    """Extract normalized Web Vocabulary class records."""
    classes: list[dict[str, Any]] = []
    for item in _graph(data):
        term_id = item.get("@id")
        if not isinstance(term_id, str) or not term_id.startswith("gs1:"):
            continue
        if not _is_class(item):
            continue
        classes.append(
            {
                "term_id": term_id,
                "compact_name": _compact_name(term_id),
                "label": _text(item.get("rdfs:label")),
                "comment": _text(item.get("rdfs:comment")),
                "sub_class_of": _values(item.get("rdfs:subClassOf")),
                "type": _types(item),
                "term_status": _text(item.get("sw:term_status")),
                "version": version,
                "last_modified": last_modified,
            }
        )
    return sorted(classes, key=lambda row: row["term_id"].lower())


def extract_webvoc_properties(
    data: dict[str, Any],
    *,
    version: str = WEBVOC_SOURCE_VERSION,
    last_modified: str = "",
    linktypes: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Extract normalized Web Vocabulary property records."""
    linktypes = linktypes or {}
    properties: list[dict[str, Any]] = []
    for item in _graph(data):
        term_id = item.get("@id")
        if not isinstance(term_id, str) or not term_id.startswith("gs1:"):
            continue
        if not _is_property(item):
            continue
        sub_property_of = _values(item.get("rdfs:subPropertyOf"))
        compact_name = _compact_name(term_id)
        properties.append(
            {
                "term_id": term_id,
                "compact_name": compact_name,
                "label": _text(item.get("rdfs:label")),
                "comment": _text(item.get("rdfs:comment")),
                "domain": _values(item.get("rdfs:domain")),
                "range": _values(item.get("rdfs:range")),
                "sub_property_of": sub_property_of,
                "type": _types(item),
                "is_link_type": (
                    "gs1:linkType" in set(sub_property_of)
                    or compact_name in linktypes
                ),
                "term_status": _text(item.get("sw:term_status")),
                "version": version,
                "last_modified": last_modified,
            }
        )
    return sorted(properties, key=lambda row: row["term_id"].lower())


def _find_source(
    manifest: dict[str, Any],
    *,
    source_id_contains: str,
    local_path: str | Path,
) -> dict[str, Any]:
    normalized_local_path = Path(local_path).as_posix()
    for source in manifest["sources"]:
        if str(source["local_path"]).replace("\\", "/") == normalized_local_path:
            return source
    for source in manifest["sources"]:
        if source_id_contains in str(source["source_id"]):
            return source
    raise ValueError(f"Manifest does not contain a source for {local_path}.")


def _duplicate_rows(
    rows: Iterable[dict[str, Any]],
    field: str,
    *,
    limit: int = 25,
) -> list[dict[str, Any]]:
    counts = Counter(str(row.get(field, "")).strip() for row in rows)
    duplicates = [
        {"value": value, "count": count}
        for value, count in counts.items()
        if value and count > 1
    ]
    return sorted(duplicates, key=lambda row: (-row["count"], row["value"]))[:limit]


def _has_value(value: Any) -> bool:
    if isinstance(value, list):
        return bool(value)
    return bool(str(value or "").strip())


def _count_with(rows: Iterable[dict[str, Any]], field: str) -> int:
    return sum(1 for row in rows if _has_value(row.get(field, "")))


def _gdsn_summary(
    result: GDSNImportResult,
    *,
    source_path: Path,
    source: dict[str, Any],
) -> dict[str, Any]:
    rows = result.rows
    attribute_rows = [row for row in rows if _row_type(row) == "attribute"]
    class_rows = [row for row in rows if _row_type(row) == "class"]
    return {
        "source_id": source["source_id"],
        "source_url": source["source_url"],
        "local_path": str(source_path).replace("\\", "/"),
        "version": source.get("version", GDSN_SOURCE_VERSION),
        "sha256": sha256_file(source_path),
        "sheet_count": result.workbook_sheet_count,
        "selected_sheet": result.selected_sheet,
        "active_rows": result.active_row_count,
        "total_rows": len(rows),
        "attribute_rows": len(attribute_rows),
        "class_rows": len(class_rows),
        "deleted_rows": result.deleted_row_count,
        "candidate_source_rows": sum(1 for row in rows if row["is_candidate_source"]),
        "rows_with_bms_id": _count_with(rows, "bms_id"),
        "rows_with_xpath": _count_with(rows, "xpath"),
        "rows_with_definition": _count_with(rows, "definition"),
        "rows_with_data_type": _count_with(rows, "data_type"),
        "rows_with_code_list": sum(
            1
            for row in rows
            if row.get("code_list_enumeration")
            or row.get("code_list_name")
            or row.get("bms_code_list_id")
        ),
        "possible_duplicate_bms_ids": _duplicate_rows(rows, "bms_id"),
        "possible_duplicate_xpath_values": _duplicate_rows(rows, "xpath"),
        "missing_critical_fields": {
            "bms_id": len(rows) - _count_with(rows, "bms_id"),
            "xpath": len(attribute_rows) - _count_with(attribute_rows, "xpath"),
            "attribute_name": len(attribute_rows)
            - _count_with(attribute_rows, "attribute_name"),
            "row_type": len(rows) - _count_with(rows, "row_type"),
        },
    }


def _webvoc_summary(
    properties: list[dict[str, Any]],
    classes: list[dict[str, Any]],
    *,
    source_path: Path,
    source: dict[str, Any],
) -> dict[str, Any]:
    return {
        "source_id": source["source_id"],
        "source_url": source["source_url"],
        "local_path": str(source_path).replace("\\", "/"),
        "version": source.get("version", WEBVOC_SOURCE_VERSION),
        "last_modified": source.get("last_modified", ""),
        "sha256": sha256_file(source_path),
        "class_count": len(classes),
        "property_count": len(properties),
        "link_type_property_count": sum(
            1 for row in properties if row.get("is_link_type")
        ),
        "stable_property_count": sum(
            1 for row in properties if row.get("term_status") == "stable"
        ),
        "properties_with_domain": _count_with(properties, "domain"),
        "properties_with_range": _count_with(properties, "range"),
        "missing_critical_fields": {
            "label": len(properties) - _count_with(properties, "label"),
            "comment": len(properties) - _count_with(properties, "comment"),
            "domain": len(properties) - _count_with(properties, "domain"),
            "range": len(properties) - _count_with(properties, "range"),
        },
    }


def build_reference_data_import(
    *,
    gdsn_xlsx: str | Path = DEFAULT_GDSN_XLSX,
    webvoc: str | Path = DEFAULT_WEBVOC,
    source_manifest: str | Path = DEFAULT_SOURCE_MANIFEST,
) -> ReferenceDataImport:
    """Build normalized reference datasets from local public source files."""
    manifest = load_source_manifest(source_manifest)
    gdsn_path = Path(gdsn_xlsx)
    webvoc_path = Path(webvoc)
    gdsn_source = _find_source(
        manifest,
        source_id_contains="gdsn",
        local_path=gdsn_path,
    )
    webvoc_source = _find_source(
        manifest,
        source_id_contains="web_vocabulary",
        local_path=webvoc_path,
    )

    checksum_checks = validate_manifest_checksums(manifest)
    mismatches = [
        check for check in checksum_checks if check["exists"] and not check["matches"]
    ]
    if mismatches:
        mismatch_text = ", ".join(check["source_id"] for check in mismatches)
        raise ValueError(f"Source manifest checksum mismatch: {mismatch_text}")

    gdsn_result = load_gdsn_bms_xpath_workbook(
        gdsn_path,
        source_version=str(gdsn_source.get("version") or GDSN_SOURCE_VERSION),
    )
    webvoc_data = load_webvoc_jsonld(webvoc_path)
    webvoc_version = str(webvoc_source.get("version") or WEBVOC_SOURCE_VERSION)
    webvoc_last_modified = str(webvoc_source.get("last_modified", ""))
    linktypes = _load_linktypes_for_webvoc(webvoc_path)
    webvoc_classes = extract_webvoc_classes(
        webvoc_data,
        version=webvoc_version,
        last_modified=webvoc_last_modified,
    )
    webvoc_properties = extract_webvoc_properties(
        webvoc_data,
        version=webvoc_version,
        last_modified=webvoc_last_modified,
        linktypes=linktypes,
    )

    summary = {
        "summary_version": "1.0",
        "offline": True,
        "sources": {
            source["source_id"]: {
                "title": source["title"],
                "source_url": source["source_url"],
                "local_path": source["local_path"],
                "version": source["version"],
                "sha256": source["sha256"],
                "public_accessible": source["public_accessible"],
                "authoritative_or_derived": source["authoritative_or_derived"],
            }
            for source in manifest["sources"]
        },
        "checksum_checks": checksum_checks,
        "gdsn": _gdsn_summary(
            gdsn_result,
            source_path=gdsn_path,
            source=gdsn_source,
        ),
        "webvoc": _webvoc_summary(
            webvoc_properties,
            webvoc_classes,
            source_path=webvoc_path,
            source=webvoc_source,
        ),
    }
    return ReferenceDataImport(
        source_manifest=manifest,
        gdsn_rows=gdsn_result.rows,
        webvoc_properties=webvoc_properties,
        webvoc_classes=webvoc_classes,
        summary=summary,
    )


def _csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, str):
        return " ".join(value.split())
    return str(value)


def _write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def _write_csv(path: Path, rows: list[dict[str, Any]], fields: tuple[str, ...]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {field: _csv_value(row.get(field, "")) for field in fields}
            )


def write_reference_data_outputs(
    import_result: ReferenceDataImport,
    output_dir: str | Path = DEFAULT_OUTPUT_DIR,
) -> dict[str, Path]:
    """Write normalized reference data JSON, CSV, and summary files."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    paths = {
        "gdsn_json": output_path / f"{GDSN_OUTPUT_BASE}.json",
        "gdsn_csv": output_path / f"{GDSN_OUTPUT_BASE}.csv",
        "webvoc_properties_json": output_path
        / f"{WEBVOC_PROPERTIES_OUTPUT_BASE}.json",
        "webvoc_properties_csv": output_path / f"{WEBVOC_PROPERTIES_OUTPUT_BASE}.csv",
        "webvoc_classes_json": output_path / f"{WEBVOC_CLASSES_OUTPUT_BASE}.json",
        "webvoc_classes_csv": output_path / f"{WEBVOC_CLASSES_OUTPUT_BASE}.csv",
        "summary_json": output_path / "source_data_summary.json",
    }
    _write_json(paths["gdsn_json"], import_result.gdsn_rows)
    _write_csv(paths["gdsn_csv"], import_result.gdsn_rows, GDSN_FIELDS)
    _write_json(paths["webvoc_properties_json"], import_result.webvoc_properties)
    _write_csv(
        paths["webvoc_properties_csv"],
        import_result.webvoc_properties,
        WEBVOC_PROPERTY_FIELDS,
    )
    _write_json(paths["webvoc_classes_json"], import_result.webvoc_classes)
    _write_csv(
        paths["webvoc_classes_csv"],
        import_result.webvoc_classes,
        WEBVOC_CLASS_FIELDS,
    )
    _write_json(paths["summary_json"], import_result.summary)
    return paths
