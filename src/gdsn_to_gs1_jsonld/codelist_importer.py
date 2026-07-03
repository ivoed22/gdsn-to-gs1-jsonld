"""Offline import for the GDSN and Shared Common Code Lists workbook (Track D).

Parses the committed public GDSN codelist export into a normalized,
deterministic, versioned registry: codelist names, their allowed values, and
their deprecated ("Deleted Codes") values with sunset release. Offline only
— no network access. Source/licensing note: see
``reference_data/source_manifest.json`` entry
``gdsn_and_shared_code_lists_r3p1p36_i6``.
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_CODELIST_XLSX = Path(
    "reference_data/raw_public/"
    "GDSN_and_Shared_Code_Lists_r3p1p36_i6_8May2026.xlsx"
)
DEFAULT_OUTPUT_DIR = Path("reference_data/normalized")
CODELIST_OUTPUT_BASE = "gdsn_codelists_r3_1_36"
CODELIST_SOURCE_VERSION = "3.1.36"

_CODELIST_SHEET = "CodeList"
_DELETED_SHEET = "Deleted Codes "  # workbook sheet name has a trailing space


@dataclass(frozen=True)
class CodelistImportResult:
    codelists: dict[str, dict[str, Any]]
    codelist_count: int
    value_count: int
    deprecated_value_count: int
    unmatched_deprecated_count: int


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return str(value)
    return str(value).replace("\r\n", "\n").strip()


def _load_codelist_sheet(workbook: Any) -> list[dict[str, str]]:
    worksheet = workbook[_CODELIST_SHEET]
    rows = worksheet.iter_rows(values_only=True)
    headers = [_cell_to_text(h) for h in next(rows)]
    records = []
    for values in rows:
        record = {
            header: _cell_to_text(value)
            for header, value in zip(headers, values)
            if header
        }
        if record.get("Code List"):
            records.append(record)
    return records


def _load_deleted_sheet(workbook: Any) -> list[dict[str, str]]:
    if _DELETED_SHEET not in workbook.sheetnames:
        return []
    worksheet = workbook[_DELETED_SHEET]
    rows = worksheet.iter_rows(values_only=True)
    headers = [_cell_to_text(h) for h in next(rows)]
    records = []
    for values in rows:
        record = {
            header: _cell_to_text(value)
            for header, value in zip(headers, values)
            if header
        }
        if record.get("Code List"):
            records.append(record)
    return records


def build_codelist_registry(
    xlsx_path: str | Path = DEFAULT_CODELIST_XLSX,
) -> CodelistImportResult:
    """Parse the codelist workbook into a normalized registry.

    Registry shape: ``{codelist_name: {"semantic_resource_urn", "domain",
    "status", "values": [{"value","label","definition","status"}],
    "deprecated_values": [{"value","label","definition","sunset_release"}]}}``.
    """
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    codelist_rows = _load_codelist_sheet(workbook)
    deleted_rows = _load_deleted_sheet(workbook)

    registry: dict[str, dict[str, Any]] = {}
    value_count = 0
    for row in codelist_rows:
        codelist_name = row["Code List"]
        entry = registry.setdefault(
            codelist_name,
            {
                "semantic_resource_urn": "",
                "domain": "",
                "status": "",
                "values": [],
                "deprecated_values": [],
            },
        )
        row_type = row.get("resourceSubTypeCode", "")
        if row_type == "CODELIST":
            entry["semantic_resource_urn"] = row.get("semanticResourceURN", "")
            entry["domain"] = row.get("Domain", "")
            entry["status"] = row.get("Status", "").strip()
        elif row_type == "CODEVALUE":
            code_value = row.get("Code Value", "")
            if not code_value:
                continue
            entry["values"].append(
                {
                    "value": code_value,
                    "label": row.get("Code Name", ""),
                    "definition": row.get("Definition", ""),
                    "status": row.get("Status", "").strip(),
                }
            )
            value_count += 1

    unmatched_deprecated = 0
    deprecated_count = 0
    for row in deleted_rows:
        codelist_name = row["Code List"]
        code_value = row.get("Code Value", "")
        if not code_value:
            continue
        if codelist_name not in registry:
            registry[codelist_name] = {
                "semantic_resource_urn": "",
                "domain": row.get("Domain", ""),
                "status": "deleted",
                "values": [],
                "deprecated_values": [],
            }
            unmatched_deprecated += 1
        registry[codelist_name]["deprecated_values"].append(
            {
                "value": code_value,
                "label": row.get("Code Name", ""),
                "definition": row.get("Definition", ""),
                "sunset_release": row.get("Sunset Release", "").strip(),
            }
        )
        deprecated_count += 1

    # Deterministic ordering.
    for entry in registry.values():
        entry["values"].sort(key=lambda item: item["value"])
        entry["deprecated_values"].sort(key=lambda item: item["value"])
    sorted_registry = {name: registry[name] for name in sorted(registry)}

    return CodelistImportResult(
        codelists=sorted_registry,
        codelist_count=len(sorted_registry),
        value_count=value_count,
        deprecated_value_count=deprecated_count,
        unmatched_deprecated_count=unmatched_deprecated,
    )


def write_codelist_registry(
    result: CodelistImportResult,
    *,
    source_path: str | Path,
    output_dir: str | Path = DEFAULT_OUTPUT_DIR,
) -> dict[str, str]:
    """Write the normalized codelist registry JSON + a summary."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    registry_path = output_path / f"{CODELIST_OUTPUT_BASE}.json"
    summary_path = output_path / f"{CODELIST_OUTPUT_BASE}_summary.json"

    payload = {
        "source_version": CODELIST_SOURCE_VERSION,
        "source_local_path": str(source_path).replace("\\", "/"),
        "source_sha256": sha256_file(source_path),
        "codelists": result.codelists,
    }
    registry_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=False) + "\n",
        encoding="utf-8",
    )
    summary = {
        "codelist_count": result.codelist_count,
        "value_count": result.value_count,
        "deprecated_value_count": result.deprecated_value_count,
        "unmatched_deprecated_count": result.unmatched_deprecated_count,
        "source_sha256": payload["source_sha256"],
    }
    summary_path.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    return {"registry": str(registry_path), "summary": str(summary_path)}
