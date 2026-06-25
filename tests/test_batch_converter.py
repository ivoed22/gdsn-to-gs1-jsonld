import json
import zipfile
from io import BytesIO

import pytest
from openpyxl import load_workbook

from gdsn_to_gs1_jsonld.batch_converter import (
    BatchConversionError,
    BatchConversionLimits,
    convert_batch_zip,
)


def _zip_bytes(entries: dict[str, bytes | str]) -> bytes:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, content in entries.items():
            data = content.encode("utf-8") if isinstance(content, str) else content
            archive.writestr(name, data)
    return buffer.getvalue()


def _zip_names(data: bytes) -> set[str]:
    with zipfile.ZipFile(BytesIO(data)) as archive:
        return set(archive.namelist())


def test_batch_zip_converts_multiple_valid_xml_files(sample_dir, mapping_v0_3_path):
    batch_zip = _zip_bytes(
        {
            "minimal_product.xml": (sample_dir / "minimal_product.xml").read_bytes(),
            "food_product_full.xml": (sample_dir / "food_product_full.xml").read_bytes(),
        }
    )

    report = convert_batch_zip(batch_zip, mapping_v0_3_path)

    assert report.xml_files_found == 2
    assert report.success_count == 2
    assert report.failure_count == 0
    assert all(file.status == "success" for file in report.files)


def test_batch_zip_ignores_non_xml_files(sample_dir, mapping_v0_3_path):
    batch_zip = _zip_bytes(
        {
            "minimal_product.xml": (sample_dir / "minimal_product.xml").read_bytes(),
            "notes.txt": "not product data",
        }
    )

    report = convert_batch_zip(batch_zip, mapping_v0_3_path)

    assert report.xml_files_found == 1
    assert report.success_count == 1
    assert "notes.txt" not in json.loads(report.summary_json_bytes)["files"][0].values()


def test_invalid_xml_does_not_stop_batch(sample_dir, mapping_v0_3_path):
    batch_zip = _zip_bytes(
        {
            "minimal_product.xml": (sample_dir / "minimal_product.xml").read_bytes(),
            "broken.xml": "<not-closed>",
        }
    )

    report = convert_batch_zip(batch_zip, mapping_v0_3_path)

    assert report.xml_files_found == 2
    assert report.success_count == 1
    assert report.failure_count == 1
    failed = next(file for file in report.files if file.status == "error")
    assert failed.original_filename == "broken.xml"
    assert failed.error_type == "XMLParseError"


def test_path_traversal_entries_are_rejected_safely(sample_dir, mapping_v0_3_path):
    batch_zip = _zip_bytes(
        {
            "../evil.xml": (sample_dir / "minimal_product.xml").read_bytes(),
            "safe/good.xml": (sample_dir / "food_product_full.xml").read_bytes(),
        }
    )

    report = convert_batch_zip(batch_zip, mapping_v0_3_path)
    names = _zip_names(report.export_zip_bytes)

    assert report.xml_files_found == 2
    assert report.success_count == 1
    assert report.failure_count == 1
    assert any(file.error_type == "UnsafeZipPath" for file in report.files)
    assert all(".." not in name for name in names)
    assert "errors/evil_error.json" in names


def test_max_file_count_limit_is_enforced(sample_dir, mapping_v0_3_path):
    batch_zip = _zip_bytes(
        {
            "one.xml": (sample_dir / "minimal_product.xml").read_bytes(),
            "two.xml": (sample_dir / "food_product_full.xml").read_bytes(),
        }
    )

    with pytest.raises(BatchConversionError, match="max_files"):
        convert_batch_zip(
            batch_zip,
            mapping_v0_3_path,
            limits=BatchConversionLimits(max_files=1),
        )


def test_max_file_size_limit_is_reported_per_file(sample_dir, mapping_v0_3_path):
    batch_zip = _zip_bytes(
        {
            "too-large.xml": (sample_dir / "minimal_product.xml").read_bytes(),
        }
    )

    report = convert_batch_zip(
        batch_zip,
        mapping_v0_3_path,
        limits=BatchConversionLimits(max_uncompressed_file_size=10),
    )

    assert report.xml_files_found == 1
    assert report.success_count == 0
    assert report.failure_count == 1
    assert report.files[0].error_type == "FileTooLarge"


def test_max_total_size_limit_is_enforced(sample_dir, mapping_v0_3_path):
    batch_zip = _zip_bytes(
        {
            "minimal_product.xml": (sample_dir / "minimal_product.xml").read_bytes(),
            "food_product_full.xml": (sample_dir / "food_product_full.xml").read_bytes(),
        }
    )

    with pytest.raises(BatchConversionError, match="max_total_uncompressed_size"):
        convert_batch_zip(
            batch_zip,
            mapping_v0_3_path,
            limits=BatchConversionLimits(max_total_uncompressed_size=10),
        )


def test_batch_summary_json_and_xlsx_are_created(
    sample_dir,
    mapping_v0_3_path,
    tmp_path,
):
    batch_zip = _zip_bytes(
        {
            "minimal_product.xml": (sample_dir / "minimal_product.xml").read_bytes(),
        }
    )

    report = convert_batch_zip(batch_zip, mapping_v0_3_path, output_dir=tmp_path)
    summary_payload = json.loads(report.summary_json_bytes)
    workbook = load_workbook(BytesIO(report.summary_xlsx_bytes), read_only=True)

    assert summary_payload["summary"]["xml_files_found"] == 1
    assert "Batch Summary" in workbook.sheetnames
    assert (tmp_path / "batch_summary.json").is_file()
    assert (tmp_path / "batch_summary.xlsx").is_file()


def test_batch_export_zip_structure_is_correct(sample_dir, mapping_v0_3_path):
    batch_zip = _zip_bytes(
        {
            "minimal_product.xml": (sample_dir / "minimal_product.xml").read_bytes(),
            "broken.xml": "<not-closed>",
        }
    )

    report = convert_batch_zip(batch_zip, mapping_v0_3_path)
    names = _zip_names(report.export_zip_bytes)

    assert "batch_summary.json" in names
    assert "batch_summary.xlsx" in names
    assert any(name.startswith("products/") and name.endswith(".jsonld") for name in names)
    assert any(
        name.startswith("reports/") and name.endswith("_mapping_report.xlsx")
        for name in names
    )
    assert any(
        name.startswith("reports/") and name.endswith("_validation_report.json")
        for name in names
    )
    assert any(
        name.startswith("reports/") and name.endswith("_unmapped_fields.json")
        for name in names
    )
    assert "errors/broken_error.json" in names
