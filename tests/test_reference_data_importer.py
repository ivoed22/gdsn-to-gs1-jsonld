import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from gdsn_to_gs1_jsonld.cli import app
from gdsn_to_gs1_jsonld.reference_data_importer import (
    GDSN_FIELDS,
    SOURCE_MANIFEST_REQUIRED_FIELDS,
    build_reference_data_import,
    extract_webvoc_classes,
    extract_webvoc_properties,
    load_gdsn_bms_xpath_workbook,
    load_source_manifest,
    load_webvoc_jsonld,
    sha256_file,
    write_reference_data_outputs,
)


ROOT = Path(__file__).resolve().parents[1]


def _create_fake_gdsn_workbook(path: Path) -> None:
    workbook = Workbook()
    change_sheet = workbook.active
    change_sheet.title = "changelist"
    change_sheet.append(["Change"])
    change_sheet.append(["Fixture"])

    active = workbook.create_sheet("3.1.36")
    headers = [
        "BMSid",
        "MESSAGE",
        "Sunrise",
        "Sunset",
        "Rel Updated",
        "ChangeType",
        "xPath",
        "Module",
        "Type",
        "_Parent_ Class",
        "Name",
        "Multiplicity",
        "Length",
        "DataType",
        "Named Association?",
        "Class Associated To",
        "CodeList_Enumeration",
        "CodeListName",
        "BMSCodeListID",
        "Notes",
        "MultipleValues",
        "LanguageEnabled",
        "MultipleLanguages",
        "UOMEnabled",
        "MultipleUOM",
        "CurrencyEnabled",
        "MultipleCurrency",
        "TPNOrTPD",
        "GlobalLocal",
        "SemanticResourceURN",
        "Definition",
    ]
    active.append(headers)
    active.append(
        [
            67,
            "CatalogueItemNotification",
            "",
            "",
            "",
            "A",
            "/tradeItem/gtin",
            "TradeItem",
            "Attribute",
            "TradeItem",
            "gtin",
            "1..1",
            "{14}",
            "string",
            "No",
            "",
            "",
            "",
            "",
            "",
            "",
            "No",
            "",
            "No",
            "",
            "No",
            "",
            "",
            "Global",
            "urn:gs1:gdsn:gtin",
            "Global Trade Item Number.",
        ]
    )
    active.append(
        [
            100,
            "CatalogueItemNotification",
            "",
            "",
            "",
            "",
            "/tradeItem",
            "TradeItem",
            "Class",
            "",
            "TradeItem",
            "1..1",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Global",
            "",
            "Trade item class.",
        ]
    )

    deleted = workbook.create_sheet("Deleted Attributes")
    deleted.append(headers)
    deleted.append(
        [
            297,
            "CatalogueItemNotification",
            "",
            "",
            "",
            "D",
            "/tradeItem/oldName",
            "TradeItem",
            "Attribute",
            "TradeItem",
            "oldName",
            "0..1",
            "{1..80}",
            "string",
            "No",
            "",
            "",
            "",
            "",
            "",
            "",
            "Yes",
            "",
            "No",
            "",
            "No",
            "",
            "",
            "Global",
            "",
            "Deleted attribute.",
        ]
    )
    workbook.save(path)


def _write_fake_webvoc(path: Path) -> None:
    data = {
        "@graph": [
            {
                "@id": "gs1:",
                "@type": ["voaf:Vocabulary", "owl:Ontology"],
                "owl:versionInfo": "1.17",
            },
            {
                "@id": "gs1:Product",
                "@type": ["owl:Class", "rdfs:Class"],
                "rdfs:label": {"@value": "Product"},
                "rdfs:comment": {"@value": "Any item offered for sale."},
                "rdfs:subClassOf": {"@id": "owl:Thing"},
                "sw:term_status": "stable",
            },
            {
                "@id": "gs1:gtin",
                "@type": ["owl:DatatypeProperty", "rdf:Property"],
                "rdfs:label": {"@value": "GTIN"},
                "rdfs:comment": {"@value": "Global Trade Item Number."},
                "rdfs:domain": {"@id": "gs1:Product"},
                "rdfs:range": {"@id": "xsd:string"},
                "sw:term_status": "stable",
            },
            {
                "@id": "gs1:linkExample",
                "@type": ["owl:ObjectProperty", "rdf:Property"],
                "rdfs:label": {"@value": "Link Example"},
                "rdfs:comment": {"@value": "Example link property."},
                "rdfs:domain": {"@id": "gs1:Product"},
                "rdfs:range": {"@id": "xsd:anyURI"},
                "rdfs:subPropertyOf": {"@id": "gs1:linkType"},
                "sw:term_status": "stable",
            },
        ]
    }
    path.write_text(json.dumps(data), encoding="utf-8-sig")


def _write_manifest(path: Path, gdsn_xlsx: Path, webvoc: Path) -> None:
    manifest = {
        "schema_version": "1.0",
        "sources": [
            {
                "source_id": "gdsn_bms_xpath_3_1_36",
                "title": "GDSN Attributes with BMSId and XPath 3.1.36",
                "source_url": "https://www.gs1.org/docs/gdsn/3.1/example.xlsx",
                "retrieved_at": "2026-06-29T00:00:00Z",
                "version": "3.1.36",
                "local_path": str(gdsn_xlsx),
                "sha256": sha256_file(gdsn_xlsx),
                "public_accessible": True,
                "authoritative_or_derived": "authoritative_public_source_copy",
                "used_by": ["test"],
                "license_or_rights_note": "Public reference material.",
                "usage_note": "Fixture workbook.",
            },
            {
                "source_id": "gs1_web_vocabulary_jsonld_1_17",
                "title": "GS1 Web Vocabulary JSON-LD",
                "source_url": "https://ref.gs1.org/voc/data/gs1Voc.jsonld",
                "retrieved_at": "2026-06-14T13:32:35Z",
                "version": "1.17",
                "last_modified": "2025-11-18",
                "local_path": str(webvoc),
                "sha256": sha256_file(webvoc),
                "public_accessible": True,
                "authoritative_or_derived": "authoritative_local_snapshot",
                "used_by": ["test"],
                "license_or_rights_note": "Public reference material.",
                "usage_note": "Fixture JSON-LD.",
            },
        ],
    }
    path.write_text(json.dumps(manifest), encoding="utf-8")


def test_sha256_file_is_deterministic(tmp_path):
    source = tmp_path / "source.txt"
    source.write_text("abc", encoding="utf-8")

    assert (
        sha256_file(source)
        == "ba7816bf8f01cfea414140de5dae2223b00361a396177a9cb410ff61f20015ad"
    )


def test_loads_utf8_sig_webvoc_and_extracts_classes_and_properties(tmp_path):
    webvoc = tmp_path / "gs1Voc.jsonld"
    _write_fake_webvoc(webvoc)

    data = load_webvoc_jsonld(webvoc)
    classes = extract_webvoc_classes(
        data,
        version="1.17",
        last_modified="2025-11-18",
    )
    properties = extract_webvoc_properties(
        data,
        version="1.17",
        last_modified="2025-11-18",
    )

    assert classes[0]["term_id"] == "gs1:Product"
    assert classes[0]["label"] == "Product"
    gtin = next(row for row in properties if row["term_id"] == "gs1:gtin")
    assert gtin["domain"] == ["gs1:Product"]
    assert gtin["range"] == ["xsd:string"]
    assert gtin["version"] == "1.17"
    link = next(row for row in properties if row["term_id"] == "gs1:linkExample")
    assert link["is_link_type"] is True


def test_gdsn_workbook_normalizes_attribute_class_deleted_and_candidate_rows(tmp_path):
    workbook = tmp_path / "gdsn.xlsx"
    _create_fake_gdsn_workbook(workbook)

    result = load_gdsn_bms_xpath_workbook(workbook)

    assert result.workbook_sheet_count == 3
    assert result.selected_sheet == "3.1.36"
    assert result.active_row_count == 2
    assert result.deleted_row_count == 1
    assert all(field in result.rows[0] for field in GDSN_FIELDS)

    attribute = next(row for row in result.rows if row["attribute_name"] == "gtin")
    assert attribute["bms_id"] == "67"
    assert attribute["xpath"] == "/tradeItem/gtin"
    assert attribute["module"] == "TradeItem"
    assert attribute["data_type"] == "string"
    assert attribute["definition"] == "Global Trade Item Number."
    assert attribute["is_candidate_source"] is True

    class_row = next(row for row in result.rows if row["row_type"] == "Class")
    assert class_row["is_candidate_source"] is False

    deleted = next(row for row in result.rows if row["attribute_name"] == "oldName")
    assert deleted["is_deleted"] is True
    assert deleted["is_candidate_source"] is False


def test_reference_import_writes_summary_and_normalized_files(tmp_path):
    workbook = tmp_path / "gdsn.xlsx"
    webvoc = tmp_path / "gs1Voc.jsonld"
    manifest = tmp_path / "source_manifest.json"
    output_dir = tmp_path / "normalized"
    _create_fake_gdsn_workbook(workbook)
    _write_fake_webvoc(webvoc)
    _write_manifest(manifest, workbook, webvoc)

    import_result = build_reference_data_import(
        gdsn_xlsx=workbook,
        webvoc=webvoc,
        source_manifest=manifest,
    )
    paths = write_reference_data_outputs(import_result, output_dir)
    summary = json.loads(paths["summary_json"].read_text(encoding="utf-8"))

    assert summary["offline"] is True
    assert summary["gdsn"]["total_rows"] == 3
    assert summary["gdsn"]["attribute_rows"] == 2
    assert summary["gdsn"]["class_rows"] == 1
    assert summary["gdsn"]["deleted_rows"] == 1
    assert summary["gdsn"]["candidate_source_rows"] == 1
    assert summary["webvoc"]["property_count"] == 2
    assert summary["webvoc"]["class_count"] == 1
    assert summary["checksum_checks"][0]["matches"] is True
    assert paths["gdsn_csv"].is_file()
    assert paths["webvoc_properties_json"].is_file()
    assert paths["webvoc_classes_csv"].is_file()


def test_import_reference_data_cli_creates_expected_files(tmp_path):
    workbook = tmp_path / "gdsn.xlsx"
    webvoc = tmp_path / "gs1Voc.jsonld"
    manifest = tmp_path / "source_manifest.json"
    output_dir = tmp_path / "normalized"
    _create_fake_gdsn_workbook(workbook)
    _write_fake_webvoc(webvoc)
    _write_manifest(manifest, workbook, webvoc)

    result = CliRunner().invoke(
        app,
        [
            "import-reference-data",
            "--gdsn-xlsx",
            str(workbook),
            "--webvoc",
            str(webvoc),
            "--source-manifest",
            str(manifest),
            "--output-dir",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    assert "Reference data import:" in result.output
    assert (output_dir / "gdsn_attributes_bms_xpath_3_1_36.csv").is_file()
    assert (output_dir / "gdsn_attributes_bms_xpath_3_1_36.json").is_file()
    assert (output_dir / "webvoc_properties_1_17.csv").is_file()
    assert (output_dir / "webvoc_properties_1_17.json").is_file()
    assert (output_dir / "webvoc_classes_1_17.csv").is_file()
    assert (output_dir / "webvoc_classes_1_17.json").is_file()
    assert (output_dir / "source_data_summary.json").is_file()


def test_committed_source_manifest_and_schema_have_required_fields():
    manifest = load_source_manifest(ROOT / "reference_data" / "source_manifest.json")
    schema = json.loads(
        (
            ROOT
            / "reference_data"
            / "schemas"
            / "source_manifest.schema.json"
        ).read_text(encoding="utf-8")
    )
    schema_required = set(schema["properties"]["sources"]["items"]["required"])

    assert len(manifest["sources"]) == 2
    assert set(SOURCE_MANIFEST_REQUIRED_FIELDS).issubset(schema_required)
    for source in manifest["sources"]:
        assert set(SOURCE_MANIFEST_REQUIRED_FIELDS).issubset(source)
        assert source["public_accessible"] is True
        assert len(source["sha256"]) == 64
